#!/usr/bin/env python3

from datetime import datetime
import os
import pathlib
import site
from pprint import pprint
import click
import openpyxl
import xlsxwriter
import xlsxwriter.exceptions

from pingintel_api.utils import pretty_time_ago

site.addsitedir("../src")

from pingintel_api import PingVisionAPIClient


""" Generate Excel to manage offline assignments for a team. """

SCRIPT_DIR = pathlib.Path(__file__).parent

api_client: PingVisionAPIClient = None


def get_team_uuid(company_name: str, team_name: str):
    ret = api_client.list_teams()

    for team in ret:
        if team["company_name"] == company_name and team["team_name"] == team_name:
            return team
    else:
        print("Team not found. Available teams:")
        for team in ret:
            print(f"* {team['company_name']} / {team['team_name']} ({team['team_uuid']})")
        raise ValueError(f"Team not found: {company_name} / {team_name}")


def get_statuses(division_uuid, status_names: list[str] | None = None):
    statuses = api_client.list_submission_statuses(division=division_uuid)
    ret = []
    for status in statuses:
        if status_names and status["name"] in status_names:
            ret.append(status)
    return ret


def get_team_members(team_uuid: str):
    members = api_client.list_team_members(team_uuid=team_uuid)
    return members


@click.group()
@click.option("-e", "--env", "--environment", "environment", default="prod", help="API environment (prod, dev, etc.)")
@click.pass_context
def cli(ctx, environment):
    global api_client
    api_client = PingVisionAPIClient(environment=environment)
    ctx.ensure_object(dict)
    ctx.obj["environment"] = environment


def write_excel_assignment_queue(
    submissions,
    scrubber_names,
    team_members,
    environment,
    company_name,
    team_name,
    output_filename="assignment_queue.xlsx",
):
    """Write the assignment queue to an Excel file using xlsxwriter.

    Args:
        submissions: The list of submission data.
        scrubber_names: List of valid scrubber names for validation.
        team_members: List of team member dicts with user_email and user_id.
        environment: The API environment (prod, dev, etc.).
        company_name: The company name.
        team_name: The team name.
        output_filename: The output Excel filename (default: assignment_queue.xlsx).
    """
    if not submissions:
        print("No submissions to write.")
        return

    workbook = xlsxwriter.Workbook(output_filename)
    worksheet = workbook.add_worksheet("Assignment Queue")

    # Write header information
    header_format = workbook.add_format({"bold": True})
    worksheet.write(0, 0, "Environment:", header_format)
    worksheet.write(0, 1, environment)
    worksheet.write(1, 0, "Company:", header_format)
    worksheet.write(1, 1, company_name)
    worksheet.write(2, 0, "Team:", header_format)
    worksheet.write(2, 1, team_name)
    worksheet.write(3, 0, "Generated:", header_format)
    worksheet.write(3, 1, datetime.now().strftime("%Y-%m-%d %I:%M %p"))

    # Start table at row 5 (0-indexed: row 5 = 6th row, leaving a blank row after header)
    table_start_row = 5

    # Create a hidden sheet for validation lists and metadata
    validation_sheet = workbook.add_worksheet("_ValidationLists")
    validation_sheet.hide()

    # Write scrubber names to column A
    for row_idx, name in enumerate(scrubber_names):
        validation_sheet.write(row_idx, 0, name)

    # Write email-to-user_id mapping to columns B and C (for assign command)
    # Column B: email, Column C: user_id
    for row_idx, member in enumerate(team_members):
        validation_sheet.write(row_idx, 1, member["user_email"])
        validation_sheet.write(row_idx, 2, member["user_id"])

    # Define a named range for the scrubber names
    if scrubber_names:
        workbook.define_name("ScrubberNames", f"='_ValidationLists'!$A$1:$A${len(scrubber_names)}")

    # Define formats
    date_format = workbook.add_format({"num_format": "yyyy-mm-dd h:mm AM/PM"})

    # Define columns to export (field_key, header_name)
    # Use None for field_key for manually populated columns
    columns = [
        ("id", "Ping ID"),
        ("workflow_status_name", "Status"),
        ("claimed_by__username", "Assigned To"),
        (None, "Review By"),
        ("insured_name", "Insured Name"),
        ("filename", "Filename"),
        ("num_buildings", "# Buildings"),
        ("data_readiness_score", "Readiness"),
        ("received_time", "Received Time"),
        ("team_name", "Team"),
        ("client_ref", "Client Ref"),
        ("inception_date", "Inception Date"),
        ("need_by_date", "Need By Date"),
        ("broker_name", "Broker Name"),
    ]

    # Find column indices for validation
    assigned_to_col = next(i for i, (key, _) in enumerate(columns) if key == "claimed_by__username")
    review_by_col = next(i for i, (key, name) in enumerate(columns) if name == "Review By")

    # Build table data
    table_data = []
    for submission in submissions:
        row_data = []
        for col_idx, (field_key, _) in enumerate(columns):
            if field_key is None:
                # Blank column (e.g., Review By)
                row_data.append("")
            elif field_key == "received_time":
                # Convert to datetime for proper Excel formatting
                value = submission.get(field_key, "")
                if value:
                    try:
                        dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
                        row_data.append(dt.replace(tzinfo=None))  # Remove timezone for Excel
                    except (ValueError, AttributeError):
                        row_data.append(value)
                else:
                    row_data.append("")
            else:
                value = submission.get(field_key, "")
                if value is None:
                    value = ""
                row_data.append(value)
        table_data.append(row_data)

    # Calculate column widths
    col_widths = []
    for col_idx, (field_key, header_name) in enumerate(columns):
        max_width = len(header_name)
        for row in table_data[:50]:  # Sample first 50 rows for width
            value = row[col_idx]
            if value:
                max_width = max(max_width, min(len(str(value)), 50))
        col_widths.append(max_width + 2)

    # Define table column options
    table_columns = []
    for col_idx, (field_key, header_name) in enumerate(columns):
        col_options = {"header": header_name}
        # Apply date format to date columns
        if field_key in ("received_time", "inception_date", "need_by_date"):
            col_options["format"] = date_format
        table_columns.append(col_options)

    # Add table with data
    num_rows = len(table_data)
    num_cols = len(columns)

    worksheet.add_table(
        table_start_row,
        0,
        table_start_row + num_rows,
        num_cols - 1,
        {
            "name": "AssignmentQueue",
            "style": "Table Style Medium 2",
            "columns": table_columns,
            "data": table_data,
        },
    )

    # Set column widths
    for col_idx, width in enumerate(col_widths):
        worksheet.set_column(col_idx, col_idx, width)

    # Overwrite Ping ID column with clickable URLs
    ping_id_col = next(i for i, (key, _) in enumerate(columns) if key == "id")
    for row_idx, submission in enumerate(submissions, start=table_start_row + 1):
        ping_id = submission.get("id", "")
        company_short_name = submission["company_short_name"]
        if ping_id:
            url = f"{api_client.api_url}/{company_short_name}/s/{ping_id}/overview"
            worksheet.write_url(row_idx, ping_id_col, url, string=ping_id)

    # Add data validation for Assigned To and Review By columns
    if num_rows > 0 and scrubber_names:
        validation_options = {
            "validate": "list",
            "source": "=ScrubberNames",
            "input_message": "Select a scrubber name",
            "error_message": "Please select a valid scrubber name from the list.",
        }
        # Validation for Assigned To column (data rows only)
        ret = worksheet.data_validation(
            table_start_row + 1, assigned_to_col, table_start_row + num_rows, assigned_to_col, validation_options
        )
        if ret < 0:
            print(f"Warning: data_validation for Assigned To failed with error code {ret}")
        # Validation for Review By column (data rows only)
        ret = worksheet.data_validation(
            table_start_row + 1, review_by_col, table_start_row + num_rows, review_by_col, validation_options
        )
        if ret < 0:
            print(f"Warning: data_validation for Review By failed with error code {ret}")

    while True:
        try:
            workbook.close()
            print(f"Excel file written: {output_filename}")
            break
        except (PermissionError, xlsxwriter.exceptions.FileCreateError):
            input(f"Cannot write to '{output_filename}' - file is open. Close it and press Enter to retry...")


@cli.command()
@click.pass_context
@click.option("--company-name", required=True, help="The team ID to view the assignment queue for.")
@click.option("--team-name", required=True, help="The team ID to view the assignment queue for.")
@click.option("--limit", type=int, default=None, help="Maximum number of items to return.")
@click.option("--auto-open", is_flag=True, default=False, help="Automatically open the output Excel file.")
def view(ctx, company_name, team_name, limit, auto_open):
    """View the current assignment queue for a team."""
    team = get_team_uuid(company_name, team_name)
    team_uuid = team["team_uuid"]
    division_uuid = team["division_uuid"]

    statuses = get_statuses(
        division_uuid=division_uuid,
        status_names=["Received", "Initializing", "Waiting for Scrubbing", "Waiting for Peer Review"],
    )
    status_uuids = [status["uuid"] for status in statuses]
    excel_output_filename = "assignment_queue.xlsx"

    items = []
    has_remaining = True
    cursor_id = None
    while has_remaining:
        queue = api_client.list_submission_activity(
            team_uuid=team_uuid,
            sort_by="created_time",
            sort_order="desc",
            page_size=200,
            workflow_status_uuid=status_uuids,
            cursor_id=cursor_id,
        )
        has_remaining = queue["has_remaining"]
        cursor_id = queue["cursor_id"]

        for submission in queue.get("results", []):
            pingid = submission["id"]
            claimed_by_username = submission["claimed_by__username"]
            claimed_by_id = submission["claimed_by_id"]
            status = submission["workflow_status_name"]

            assignee = f"{claimed_by_username} ({claimed_by_id})" if claimed_by_username else "Unassigned"

            created_time = submission["created_time"]
            age_str = pretty_time_ago(datetime.fromisoformat(created_time))

            docs = submission.get("documents", [])
            for doc in docs:
                if doc["document_type"] in ("SOV", "ACORD"):
                    filename = doc["filename"]
                    break
            insured_name = submission.get("insured_name", "N/A")
            submission["filename"] = filename

            print(f"* {pingid}, {assignee}, {age_str}, {status}, {insured_name}, {filename}")
            items.append(submission)

            # Check if we've reached the max items limit
            if limit is not None and len(items) >= limit:
                has_remaining = False
                break

    team_members = get_team_members(team_uuid=team_uuid)
    scrubber_names = [
        member["user_email"]
        for member in team_members
        if member["user_email"].endswith("@pingintel.com") or member["user_email"].endswith("@dha-1.com")
    ]

    environment = ctx.obj["environment"]
    write_excel_assignment_queue(
        items, scrubber_names, team_members, environment, company_name, team_name, excel_output_filename
    )

    if auto_open:
        os.startfile(excel_output_filename)


@cli.command()
@click.option("--input-file", default="assignment_queue.xlsx", help="The Excel file to read assignments from.")
@click.option("--dry-run", is_flag=True, default=False, help="Show what would be assigned without making changes.")
def assign(input_file, dry_run):
    """Read Excel file and assign submissions based on the 'Assigned To' column."""
    # Load the Excel file
    try:
        wb = openpyxl.load_workbook(input_file)
    except FileNotFoundError:
        print(f"Error: File '{input_file}' not found.")
        return

    ws = wb.active

    # Read metadata from the header rows
    environment = ws.cell(row=1, column=2).value
    company_name = ws.cell(row=2, column=2).value
    team_name = ws.cell(row=3, column=2).value

    if not environment or not company_name or not team_name:
        print("Error: Could not read environment/company/team from Excel header.")
        print("Make sure this file was generated by the 'view' command.")
        return

    print(f"Reading assignments from: {input_file}")
    print(f"Environment: {environment}, Company: {company_name}, Team: {team_name}")

    # Read email-to-user_id mapping from the hidden validation sheet
    if "_ValidationLists" not in wb.sheetnames:
        print("Error: Could not find validation data in Excel file.")
        print("Make sure this file was generated by the 'view' command.")
        return

    validation_ws = wb["_ValidationLists"]
    email_to_user_id = {}
    for row in validation_ws.iter_rows(min_row=1, values_only=True):
        # Column B (index 1) = email, Column C (index 2) = user_id
        if row[1] and row[2]:
            email_to_user_id[row[1]] = row[2]

    if not email_to_user_id:
        print("Error: No user mapping found in Excel file.")
        return

    # Find the table header row (row 6, 1-indexed)
    table_header_row = 6
    headers = {cell.value: idx for idx, cell in enumerate(ws[table_header_row])}

    ping_id_col = headers.get("Ping ID")
    assigned_to_col = headers.get("Assigned To")

    if ping_id_col is None or assigned_to_col is None:
        print("Error: Could not find 'Ping ID' or 'Assigned To' columns in the Excel file.")
        return

    # Process each row and collect assignments
    assignments = []  # List of (ping_id, assigned_to_email, user_id)
    assignments_skipped = 0

    for row in ws.iter_rows(min_row=table_header_row + 1, values_only=True):
        ping_id = row[ping_id_col]
        assigned_to = row[assigned_to_col]

        if not ping_id:
            continue

        if not assigned_to:
            assignments_skipped += 1
            continue

        # Look up the user ID from email
        user_id = email_to_user_id.get(assigned_to)
        if not user_id:
            print(f"Warning: Unknown assignee '{assigned_to}' for {ping_id} - skipping")
            assignments_skipped += 1
            continue

        assignments.append((ping_id, assigned_to, user_id))

    if not assignments:
        print("No assignments to make.")
        return

    if dry_run:
        for ping_id, assigned_to, user_id in assignments:
            print(f"[DRY RUN] Would assign {ping_id} to {assigned_to}")
        print(f"\nDone. Assignments would be made: {len(assignments)}, Skipped: {assignments_skipped}")
        return

    # Group assignments by user_id for bulk update
    from collections import defaultdict

    assignments_by_user = defaultdict(list)
    for ping_id, assigned_to, user_id in assignments:
        assignments_by_user[user_id].append((ping_id, assigned_to))

    assignments_made = 0
    assignments_failed = 0

    for user_id, user_assignments in assignments_by_user.items():
        ping_ids = [pa[0] for pa in user_assignments]
        assigned_to = user_assignments[0][1]  # All have same assignee

        print(f"Assigning {len(ping_ids)} submissions to {assigned_to}...")

        try:
            changes = [{"action": "claim", "parameters": {"claimed_by_id": user_id}}]
            ret = api_client.bulk_update_submission(pingids=ping_ids, changes=changes)
            # ret looks like this
            # {'results': [{'id': 'p-ob-amw-agtdgc', 'updated_data': {'claimed_by_id': 34}}]}

            for result in ret["results"]:
                if "error" in result:
                    print(f"  Failed to assign {result['id']}: {result['error']}")
                    assignments_failed += 1
                else:
                    print(f"  Assigned {result['id']} to {assigned_to}")
                    assignments_made += 1
        except Exception as e:
            print(f"  Error: {e}")
            assignments_failed += len(ping_ids)

    print(f"\nDone. Assignments made: {assignments_made}, Failed: {assignments_failed}, Skipped: {assignments_skipped}")


if __name__ == "__main__":
    cli()
